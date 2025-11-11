from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

from .forms import ClinicalSearchForm
from .permissions import check_researcher_access
from .utils import perform_fulltext_search_queryset

@login_required
def clinical_search(request):
    """
    Clinical research full text search view with Django pagination.
    """
    # Check researcher permissions
    try:
        check_researcher_access(request.user)
    except PermissionDenied as e:
        messages.error(request, str(e))
        return render(request, 'research/access_denied.html')

    form = ClinicalSearchForm()
    patients = None
    query = None

    if request.method == 'POST' or request.GET.get('q'):
        # Handle both POST (new search) and GET (pagination)
        if request.method == 'POST':
            form = ClinicalSearchForm(request.POST)
        else:
            # For pagination, get query from URL parameter
            query = request.GET.get('q', '')
            form = ClinicalSearchForm({'query': query})
        
        if form.is_valid():
            query = form.cleaned_data['query']

            # Get the queryset of patient results (limit to 100 for UI)
            try:
                patients = perform_fulltext_search_queryset(query, max_patients=100)
                if not patients:
                    messages.info(request, f'Nenhum resultado encontrado para "{query}"')
            except Exception as e:
                messages.error(request, f'Erro na busca: {str(e)}')
                patients = None

    # Apply pagination using Django's Paginator
    page_obj = None
    if patients is not None:
        paginator = Paginator(patients, 10)  # 10 results per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

    context = {
        'form': form,
        'patients': page_obj.object_list if page_obj else None,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages() if page_obj else False,
        'query': query,
        'page_title': 'Pesquisa Clínica'
    }

    return render(request, 'research/search.html', context)

@login_required
def search_ajax(request):
    """
    AJAX endpoint for search (optional for future enhancement).
    """
    try:
        check_researcher_access(request.user)
    except PermissionDenied:
        return JsonResponse({'error': 'Access denied'}, status=403)

    query = request.GET.get('q', '').strip()
    page = int(request.GET.get('page', 1))

    if not query:
        return JsonResponse({'error': 'Query required'}, status=400)

    results = perform_fulltext_search(query, page=page)

    # Convert results to JSON-serializable format
    if 'patients' in results:
        for patient_result in results['patients']:
            patient_result['patient'] = {
                'id': str(patient_result['patient'].pk),
                'name': patient_result['patient'].name
            }
            for match in patient_result['matching_notes']:
                match['note_date'] = match['note_date'].isoformat()

    return JsonResponse(results)

@login_required
def export_search_results(request):
    """
    Export search results to Excel format.
    """
    try:
        check_researcher_access(request.user)
    except PermissionDenied:
        messages.error(request, 'Acesso negado para exportação.')
        return redirect('apps.research:clinical_search')

    query = request.GET.get('q', '').strip()
    if not query:
        messages.error(request, 'Consulta de busca é obrigatória para exportação.')
        return redirect('apps.research:clinical_search')

    try:
        # Get search results for export (limit to 500 most relevant patients)
        patients = perform_fulltext_search_queryset(query, max_patients=500)
        if not patients:
            messages.info(request, 'Nenhum resultado encontrado para exportar.')
            return redirect('apps.research:clinical_search')
        
        # Warn if hitting the limit
        if len(patients) == 500:
            messages.warning(request, "Exportação limitada aos 500 pacientes mais relevantes.")

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados da Busca"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'Prontuário',
            'Iniciais',
            'Sexo',
            'Data de Nascimento',
            'Total de Resultados',
            'Melhor Relevância',
            'Trechos Encontrados'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, patient_result in enumerate(patients, 2):
            ws.cell(row=row_num, column=1, value=patient_result['registration_number'])
            ws.cell(row=row_num, column=2, value=patient_result['initials'])
            ws.cell(row=row_num, column=3, value=patient_result['gender'])
            
            # Format birthday as DD/MM/YYYY
            birthday = patient_result['birthday']
            if birthday:
                formatted_birthday = birthday.strftime('%d/%m/%Y')
            else:
                formatted_birthday = ''
            ws.cell(row=row_num, column=4, value=formatted_birthday)
            
            ws.cell(row=row_num, column=5, value=patient_result['total_matches'])
            ws.cell(row=row_num, column=6, value=round(patient_result['highest_rank'], 4))
            
            # Combine all matching notes into one cell (keep HTML tags for Excel)
            snippets = []
            matching_notes = patient_result.get('matching_notes', [])
            for match in matching_notes:
                try:
                    # Keep HTML tags - Excel will display them as text which is fine
                    if isinstance(match, dict) and 'note_date' in match and 'headline' in match:
                        snippet = f"[{match['note_date'].strftime('%d/%m/%Y %H:%M')}] {match['headline']}"
                        snippets.append(snippet)
                    else:
                        # Handle unexpected data format
                        snippets.append(f"[Data malformada: {str(match)[:100]}]")
                except Exception as e:
                    # Handle any date formatting or other errors
                    snippets.append(f"[Erro ao processar resultado: {str(e)}]")
            
            ws.cell(row=row_num, column=7, value='\n\n'.join(snippets))

        # Set fixed column widths (much faster than auto-calculation)
        column_widths = [15, 12, 10, 12, 15, 12, 60]  # Fixed widths for each column
        for col, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'busca_clinica_{query[:20]}_{timestamp}.xlsx'

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        messages.error(request, f'Erro na exportação: {str(e)}')
        return redirect('apps.research:clinical_search')
